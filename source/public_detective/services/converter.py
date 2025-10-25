"""This module provides a service for converting files."""

import csv
import io
import os
import tempfile
import zipfile

import imageio
import mammoth
import openpyxl
import textract
import xlrd
from openpyxl.worksheet.worksheet import Worksheet
from PIL import Image
from public_detective.constants.analysis_feedback import Warnings
from public_detective.providers.logging import Logger, LoggingProvider
from pyxlsb import open_workbook as open_xlsb
from striprtf.striprtf import rtf_to_text


class ConverterService:
    """A service for converting various file types for AI analysis."""

    def __init__(self) -> None:
        """Initializes the service."""
        self.logger: Logger = LoggingProvider().get_logger()

    def gif_to_mp4(self, gif_content: bytes) -> bytes:
        """Converts a GIF file content to an MP4 file content.

        Args:
            gif_content: The content of the GIF file.

        Returns:
            The content of the converted MP4 file.
        """
        self.logger.info("Converting GIF to MP4.")
        try:
            reader = imageio.get_reader(gif_content, format="gif")  # type: ignore[arg-type]
            fps = reader.get_meta_data().get("fps", 24)

            output_buffer = io.BytesIO()
            with imageio.get_writer(output_buffer, format="mp4", fps=fps) as writer:  # type: ignore[arg-type]
                for frame in reader:  # type: ignore[attr-defined]
                    writer.append_data(frame)  # type: ignore[attr-defined]

            return output_buffer.getvalue()
        except Exception as e:
            self.logger.error(f"GIF to MP4 conversion failed: {e}", exc_info=True)
            raise

    def bmp_to_png(self, bmp_content: bytes) -> bytes:
        """Converts a BMP file content to a PNG file content.

        Args:
            bmp_content: The content of the BMP file.

        Returns:
            The content of the converted PNG file.
        """
        self.logger.info("Converting BMP to PNG.")
        try:
            with Image.open(io.BytesIO(bmp_content)) as img:
                with io.BytesIO() as output_buffer:
                    img.save(output_buffer, format="PNG")
                    return output_buffer.getvalue()
        except Exception as e:
            self.logger.error(f"BMP to PNG conversion failed: {e}", exc_info=True)
            raise

    def docx_to_html(self, docx_content: bytes) -> str:
        """Converts a DOCX file content to an HTML string.

        Args:
            docx_content: The content of the DOCX file.

        Returns:
            The content of the converted HTML as a string.
        """
        self.logger.info("Converting DOCX to HTML.")
        try:
            docx_file = io.BytesIO(docx_content)
            result = mammoth.convert_to_html(docx_file)
            return str(result.value)
        except zipfile.BadZipFile:
            self.logger.warning("Mammoth conversion failed, attempting fallback with textract.")
            with tempfile.NamedTemporaryFile(delete=False, suffix=".doc") as temp_file:
                temp_file.write(docx_content)
                temp_file_path = temp_file.name

            try:
                text = textract.process(temp_file_path).decode("utf-8")
            finally:
                os.remove(temp_file_path)
            return str(text)

    def rtf_to_text(self, rtf_content: bytes) -> str:
        """Converts an RTF file content to a plain text string.

        Args:
            rtf_content: The content of the RTF file.

        Returns:
            The content of the converted text as a string.
        """
        self.logger.info("Converting RTF to text.")
        return str(rtf_to_text(rtf_content.decode("ascii", errors="ignore")))

    def doc_to_text(self, doc_content: bytes) -> str:
        """Converts a DOC file content to a plain text string.

        Args:
            doc_content: The content of the DOC file.

        Returns:
            The content of the converted text as a string.
        """
        self.logger.info("Converting DOC to text.")
        with tempfile.NamedTemporaryFile(delete=False, suffix=".doc") as temp_file:
            temp_file.write(doc_content)
            temp_file_path = temp_file.name

        text = textract.process(temp_file_path).decode("utf-8")

        os.remove(temp_file_path)

        return str(text)

    def spreadsheet_to_csvs(
        self, xls_content: bytes, original_extension: str
    ) -> tuple[list[tuple[str, bytes]], list[str]]:
        """Converts an XLS, XLSX, or XLSB file to one or more CSV files (one per sheet).

        Args:
            xls_content: The content of the spreadsheet file.
            original_extension: The original extension of the file.

        Returns:
            A tuple containing a list of tuples (sheet name, CSV content) and a list of warnings.
        """
        self.logger.info(f"Converting {original_extension} to CSV(s).")
        output_files = []
        warnings = []
        try:
            if original_extension == ".xls":
                workbook = xlrd.open_workbook(file_contents=xls_content)
                for sheet_name in workbook.sheet_names():
                    sheet = workbook.sheet_by_name(sheet_name)
                    output = io.StringIO()
                    writer = csv.writer(output)
                    for row_idx in range(sheet.nrows):
                        writer.writerow(sheet.row_values(row_idx))
                    csv_content = output.getvalue().encode("utf-8")
                    output_files.append((f"{sheet_name}.csv", csv_content))
            elif original_extension == ".xlsx":
                workbook = openpyxl.load_workbook(io.BytesIO(xls_content))
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    if not isinstance(sheet, Worksheet):
                        warnings.append(
                            Warnings.IGNORED_NON_DATA_SHEET.format(
                                sheet_name=sheet_name, sheet_type=type(sheet).__name__
                            )
                        )
                        continue
                    output = io.StringIO()
                    writer = csv.writer(output)
                    for row in sheet.iter_rows():
                        writer.writerow([cell.value for cell in row])
                    csv_content = output.getvalue().encode("utf-8")
                    output_files.append((f"{sheet_name}.csv", csv_content))
            elif original_extension == ".xlsb":
                with open_xlsb(io.BytesIO(xls_content)) as workbook:
                    for sheet_name in workbook.sheets:
                        sheet = workbook.get_sheet(sheet_name)
                        output = io.StringIO()
                        writer = csv.writer(output)
                        for row in sheet.rows():
                            writer.writerow([cell.v for cell in row])
                        csv_content = output.getvalue().encode("utf-8")
                        output_files.append((f"{sheet_name}.csv", csv_content))
            return output_files, warnings
        except Exception as e:
            self.logger.error(f"Spreadsheet conversion failed for {original_extension}: {e}", exc_info=True)
            raise
