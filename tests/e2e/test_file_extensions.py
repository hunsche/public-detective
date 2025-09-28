"This module contains E2E tests for file extension handling."

import json
import os
import shutil
import uuid
from collections.abc import Callable
from datetime import datetime
from pathlib import Path
from typing import Any

import pytest
from docx import Document
from openpyxl import Workbook
from PIL import Image
from public_detective.models.procurement_analysis_status import ProcurementAnalysisStatus
from public_detective.models.procurements import Procurement
from public_detective.providers.config import ConfigProvider
from public_detective.providers.gcs import GcsProvider
from public_detective.providers.pubsub import PubSubProvider
from public_detective.repositories.procurements import ProcurementsRepository
from public_detective.services.analysis import AnalysisService
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from sqlalchemy import text
from sqlalchemy.engine import Engine

from tests.e2e.conftest import MockPNCP, run_command


# Helper functions to generate files
def create_txt(path: Path, content: str = "This is a test text file.") -> None:
    """Creates a simple text file."""
    path.write_text(content)


def create_doc(path: Path) -> None:
    """Copies a valid DOC file from the fixtures directory."""
    fixture_path = Path(__file__).parent.parent / "fixtures" / "file_samples" / "valid_test.doc"
    shutil.copy(fixture_path, path)


def create_docx(path: Path, content: str = "This is a test DOCX file.") -> None:
    """Creates a simple DOCX file."""
    document = Document()
    document.add_paragraph(content)
    document.save(str(path))


def create_xls(path: Path) -> None:
    """Creates a simple XLS file (by saving as XLSX)."""
    workbook = Workbook()
    sheet = workbook.active
    if sheet:
        sheet["A1"] = "This is a test XLS file."
    workbook.save(str(path))


def create_xlsx(path: Path) -> None:
    """Creates a simple XLSX file."""
    workbook = Workbook()
    sheet = workbook.active
    if sheet:
        sheet.title = "Sheet1"
        sheet["A1"] = "This is a test XLSX file on Sheet1."
    sheet2 = workbook.create_sheet(title="Sheet2")
    sheet2["A1"] = "This is a test XLSX file on Sheet2."
    workbook.save(str(path))


def create_jpg(path: Path) -> None:
    """Creates a simple JPG image."""
    img = Image.new("RGB", (10, 10), color="red")
    img.save(path, "jpeg")


def create_png(path: Path) -> None:
    """Creates a simple PNG image."""
    img = Image.new("RGB", (10, 10), color="green")
    img.save(path, "png")


def create_gif(path: Path) -> None:
    """Creates a simple GIF image."""
    img = Image.new("RGB", (10, 10), color="blue")
    img.save(path, "gif")


def create_bmp(path: Path) -> None:
    """Creates a simple BMP image."""
    img = Image.new("RGB", (10, 10), color="yellow")
    img.save(path, "bmp")


def create_pdf(path: Path, content: str = "This is a valid PDF file.") -> None:
    """Creates a simple, valid PDF file."""
    pdf_canvas = canvas.Canvas(str(path), pagesize=letter)
    pdf_canvas.drawString(100, 750, content)
    pdf_canvas.save()


def create_html(path: Path, content: str = "<h1>This is a test HTML file.</h1>") -> None:
    """Creates a simple HTML file."""
    path.write_text(content)


def create_rtf(path: Path, content: str = "This is a test RTF file.") -> None:
    """Creates a simple RTF file."""
    rtf_content = f"{{{{\rtf1\\ansi\\deff0 {{{{\fonttbl {{{{\f0 Arial;}}}}}}}}\\f0\\fs24 {content}}}}}"
    path.write_text(rtf_content)


def create_xlsb(path: Path) -> None:
    """Copies a valid XLSB file from the fixtures directory."""
    fixture_path = Path(__file__).parent.parent / "fixtures" / "file_samples" / "valid_test.xlsb"
    shutil.copy(fixture_path, path)


def create_csv(path: Path, content: str = "col1,col2\nval1,val2") -> None:
    """Creates a simple CSV file."""
    path.write_text(content)


def create_json(path: Path, content: str = """{"key": "value"}""") -> None:
    """Creates a simple JSON file."""
    path.write_text(content)


def create_md(path: Path, content: str = "# Markdown") -> None:
    """Creates a simple Markdown file."""
    path.write_text(content)


def create_xml(path: Path, content: str = "<root><test>value</test></root>") -> None:
    """Creates a simple XML file."""
    path.write_text(content)


def create_media_from_fixture(path: Path, source_filename: str) -> None:
    """Copies a valid media file from the fixtures directory."""
    fixture_path = Path(__file__).parent.parent / "fixtures" / "file_samples" / source_filename
    shutil.copy(fixture_path, path)


FILE_GENERATORS: dict[str, Callable[..., None]] = {
    ".txt": create_txt,
    ".pdf": create_pdf,
    ".doc": create_doc,
    ".docx": create_docx,
    ".xls": create_xls,
    ".xlsx": create_xlsx,
    ".xlsb": create_xlsb,
    ".jpg": create_jpg,
    ".jpeg": create_jpg,  # Use the same generator for jpg and jpeg
    ".png": create_png,
    ".gif": create_gif,
    ".bmp": create_bmp,
    ".html": create_html,
    ".rtf": create_rtf,
    ".csv": create_csv,
    ".json": create_json,
    ".md": create_md,
    ".mp4": lambda p: create_media_from_fixture(p, "valid_test.mp4"),
    ".mov": lambda p: create_media_from_fixture(p, "valid_test.mov"),
    ".avi": lambda p: create_media_from_fixture(p, "valid_test.avi"),
    ".mkv": lambda p: create_media_from_fixture(p, "valid_test.mkv"),
    ".mp3": lambda p: create_media_from_fixture(p, "valid_test.mp3"),
    ".wav": lambda p: create_media_from_fixture(p, "valid_test.wav"),
    ".flac": lambda p: create_media_from_fixture(p, "valid_test.flac"),
    ".ogg": lambda p: create_media_from_fixture(p, "valid_test.ogg"),
    ".xml": create_xml,
}

SUPPORTED_EXTENSIONS_PARAMS = sorted(set(AnalysisService._SUPPORTED_EXTENSIONS))


@pytest.mark.parametrize("extension", SUPPORTED_EXTENSIONS_PARAMS)
def test_file_extension_processing(
    db_session: Engine,
    e2e_pubsub: tuple[Any, Any],
    tmp_path: Path,
    extension: str,
    mock_pncp_server: MockPNCP,
    gcs_cleanup_manager: Callable[[str], None],
) -> None:
    """
    Tests the full E2E processing for various file extensions by running the
    worker as a subprocess and using a mock PNCP server to provide the files.
    """
    expected_status = ProcurementAnalysisStatus.ANALYSIS_SUCCESSFUL

    publisher, topic_path = e2e_pubsub
    gcs_provider = GcsProvider()
    config = ConfigProvider.get_config()
    bucket_name = config.GCP_GCS_BUCKET_PROCUREMENTS
    gcs_client = gcs_provider.get_client()

    # 1. Generate file locally
    file_generator = FILE_GENERATORS[extension]
    file_name = f"test_document{extension}"
    local_file_path = tmp_path / file_name
    file_generator(local_file_path)
    assert local_file_path.exists()

    # 2. Configure mock PNCP server
    file_id = uuid.uuid4()
    procurement_control_number = f"file-ext-test-{uuid.uuid4().hex[:6]}"
    gcs_cleanup_manager(procurement_control_number)

    mock_pncp_server.file_content = local_file_path.read_bytes()
    mock_pncp_server.file_metadata = [  # type: ignore
        {
            "id": str(file_id),
            "url": f"{mock_pncp_server.url}/pncp-api/v1/contratacoes/{file_id}/arquivos/{file_name}",
            "titulo": file_name,
            "tipoDocumentoId": 2,
            "ativo": True,
            "sequencialDocumento": 1,
            "dataPublicacaoPncp": datetime.now().isoformat(),
            "cnpj": "00000000000191",
            "anoCompra": 2025,
            "sequencialCompra": 1,
            "statusAtivo": True,
            "tipoDocumentoNome": "Edital de Convocação",
            "tipoDocumentoDescricao": "Edital de Convocação",
        }
    ]
    os.environ["PNCP_INTEGRATION_API_URL"] = mock_pncp_server.url

    # 3. Setup database records
    analysis_id = uuid.uuid4()
    version_number = 1

    procurement_repo = ProcurementsRepository(engine=db_session, pubsub_provider=PubSubProvider())
    raw_data_json = json.dumps(
        {
            "anoCompra": 2025,
            "dataAtualizacao": "2025-08-23T14:30:00",
            "dataPublicacaoPncp": "2025-08-23T14:30:00",
            "sequencialCompra": 1,
            "numeroControlePNCP": procurement_control_number,
            "objetoCompra": f"Test for {extension}",
            "srp": False,
            "orgaoEntidade": {
                "cnpj": "00000000000191",
                "razaoSocial": "Test Org",
                "poderId": "E",
                "esferaId": "F",
            },
            "processo": "1/2025",
            "amparoLegal": {"codigo": 1, "nome": "Lei 14.133/2021", "descricao": "Art. 75, II"},
            "numeroCompra": "1/2025",
            "unidadeOrgao": {
                "codigoUnidade": "1",
                "nomeUnidade": "Test Unit",
                "ufNome": "SP",
                "ufSigla": "SP",
                "municipioNome": "SAO PAULO",
                "codigoIbge": "3550308",
            },
            "modalidadeId": 1,
            "dataAtualizacaoGlobal": "2025-08-23T14:30:00",
            "modoDisputaId": 1,
            "situacaoCompraId": 1,
            "usuarioNome": "Test",
        }
    )
    procurement_model = Procurement.model_validate(json.loads(raw_data_json))
    procurement_repo.save_procurement_version(
        procurement=procurement_model,
        raw_data=raw_data_json,
        version_number=version_number,
        content_hash=f"dummy_hash_{analysis_id}",
    )

    with db_session.connect() as connection:
        connection.execute(
            text(
                """INSERT INTO procurement_analyses (
                    analysis_id, procurement_control_number, version_number, status, analysis_prompt
                )
                   VALUES (
                    :analysis_id, :procurement_control_number, :version_number, 'PENDING_ANALYSIS', ''
                )"""
            ),
            {
                "analysis_id": analysis_id,
                "procurement_control_number": procurement_control_number,
                "version_number": version_number,
            },
        )
        connection.commit()
    # 4. Trigger worker by publishing a message
    message_data = {"analysis_id": str(analysis_id)}
    message_json = json.dumps(message_data)
    publisher.publish(topic_path, message_json.encode())
    print(f"Published message for analysis_id: {analysis_id}")

    # 5. Run the worker as a subprocess
    worker_command = "poetry run python -m public_detective.worker --max-messages 1 --timeout 15"
    run_command(worker_command)

    # 6. Assertions
    with db_session.connect() as connection:
        final_status = connection.execute(
            text("SELECT status FROM procurement_analyses WHERE analysis_id = :analysis_id"),
            {"analysis_id": analysis_id},
        ).scalar_one_or_none()

        assert (
            final_status == expected_status.value
        ), f"Expected {expected_status.value} for {extension}, but got {final_status}"

        file_record = (
            connection.execute(
                text(
                    """
                    SELECT fr.*
                    FROM file_records fr
                    JOIN procurement_source_documents psd ON fr.source_document_id = psd.id
                    WHERE psd.analysis_id = :analysis_id AND fr.file_name = :file_name
                    """
                ),
                {"analysis_id": analysis_id, "file_name": file_name},
            )
            .mappings()
            .one()
        )

        assert (
            file_record["exclusion_reason"] is None
        ), f"File was unexpectedly excluded with reason: {file_record['exclusion_reason']}"
        assert file_record["included_in_analysis"] is True, "File was not included in analysis"

        assert gcs_client.bucket(bucket_name).blob(file_record["gcs_path"]).exists()
