"""Unit tests for the ConverterService."""

from unittest.mock import MagicMock, patch

import pytest
from openpyxl.worksheet.worksheet import Worksheet
from public_detective.services.converter import ConverterService


@pytest.fixture
def converter_service() -> ConverterService:
    """Provides a ConverterService instance."""
    return ConverterService()


def test_rtf_to_text() -> None:
    """Test the rtf_to_text method."""
    service = ConverterService()
    result = service.rtf_to_text(b"{\\rtf1\\ansi...}")
    assert isinstance(result, str)


def test_doc_to_text() -> None:
    """Test the doc_to_text method."""
    service = ConverterService()
    with patch("textract.process") as mock_process:
        mock_process.return_value = b"text content"
        result = service.doc_to_text(b"dummy doc content")
        assert result == "text content"


def test_docx_to_html() -> None:
    """Test the docx_to_html method."""
    service = ConverterService()
    with patch("mammoth.convert") as mock_convert:
        mock_convert.return_value.value = "<html></html>"
        result = service.docx_to_html(b"dummy docx content")
        assert result == "<html></html>"


def test_gif_to_mp4() -> None:
    """Test the gif_to_mp4 method."""
    service = ConverterService()
    dummy_gif_content = b"GIF89a..."
    with patch("imageio.get_reader") as mock_get_reader, patch("imageio.get_writer") as mock_get_writer:

        mock_reader = mock_get_reader.return_value
        mock_reader.get_meta_data.return_value = {"fps": 30}
        mock_reader.__iter__.return_value = [b"frame1", b"frame2"]

        mock_writer_instance = mock_get_writer.return_value.__enter__.return_value

        result = service.gif_to_mp4(dummy_gif_content)

        mock_get_reader.assert_called_once_with(dummy_gif_content, format="gif")
        mock_get_writer.assert_called_once()
        assert mock_writer_instance.append_data.call_count == 2
        assert isinstance(result, bytes)


def test_bmp_to_png() -> None:
    """Test the bmp_to_png method."""
    service = ConverterService()
    dummy_bmp_content = (
        b"BM\x1e\x00\x00\x00\x00\x00\x00\x00\x1a\x00\x00\x00\x0c\x00\x00\x00"
        b"\x01\x00\x01\x00\x01\x00\x18\x00\x00\x00\xff\x00"
    )
    with patch("PIL.Image.open") as mock_open:
        mock_img = MagicMock()
        mock_open.return_value.__enter__.return_value = mock_img
        result = service.bmp_to_png(dummy_bmp_content)
        mock_img.save.assert_called_once()
        assert isinstance(result, bytes)


def test_spreadsheet_to_csvs_xls() -> None:
    """Test the spreadsheet_to_csvs method for XLS files."""
    service = ConverterService()
    dummy_xls_content = b"dummy xls content"
    with patch("xlrd.open_workbook") as mock_open_workbook:
        mock_workbook = mock_open_workbook.return_value
        mock_sheet = mock_workbook.sheet_by_name.return_value
        mock_sheet.nrows = 1
        mock_sheet.row_values.return_value = ["a", "b"]
        mock_workbook.sheet_names.return_value = ["sheet1"]
        files, warnings = service.spreadsheet_to_csvs(dummy_xls_content, ".xls")
        assert len(files) == 1
        assert files[0][0] == "sheet1.csv"
        assert files[0][1] == b"a,b\r\n"
        assert len(warnings) == 0


def test_spreadsheet_to_csvs_xlsx() -> None:
    """Test the spreadsheet_to_csvs method for XLSX files."""
    service = ConverterService()
    dummy_xlsx_content = b"dummy xlsx content"
    with patch("openpyxl.load_workbook") as mock_load_workbook:
        mock_workbook = mock_load_workbook.return_value
        mock_sheet = MagicMock(spec=Worksheet)
        mock_sheet.iter_rows.return_value = [[MagicMock(value="a"), MagicMock(value="b")]]
        mock_workbook.__getitem__.return_value = mock_sheet
        mock_workbook.sheetnames = ["sheet1"]
        files, warnings = service.spreadsheet_to_csvs(dummy_xlsx_content, ".xlsx")
        assert len(files) == 1
        assert files[0][0] == "sheet1.csv"
        assert files[0][1] == b"a,b\r\n"
        assert len(warnings) == 0


def test_spreadsheet_to_csvs_xlsb() -> None:
    """Test the spreadsheet_to_csvs method for XLSB files."""
    service = ConverterService()
    with open("tests/fixtures/file_samples/valid_test.xlsb", "rb") as f:
        xlsb_content = f.read()
    files, warnings = service.spreadsheet_to_csvs(xlsb_content, ".xlsb")
    assert len(files) == 3
    assert files[0][0] == "Sheet1.csv"
    assert files[0][1].startswith(b"Welcome to File Extension FYI Center!")
    assert len(warnings) == 0


@patch("imageio.get_reader")
def test_gif_to_mp4_conversion_failure(mock_get_reader: MagicMock, converter_service: ConverterService) -> None:
    """Tests that an exception during GIF to MP4 conversion is handled."""
    mock_get_reader.side_effect = Exception("imageio error")
    with pytest.raises(Exception, match="imageio error"):
        converter_service.gif_to_mp4(b"bad gif content")


@patch("PIL.Image.open")
def test_bmp_to_png_conversion_failure(mock_image_open: MagicMock, converter_service: ConverterService) -> None:
    """Tests that an exception during BMP to PNG conversion is handled."""
    mock_image_open.side_effect = Exception("PIL error")
    with pytest.raises(Exception, match="PIL error"):
        converter_service.bmp_to_png(b"bad bmp content")


@patch("xlrd.open_workbook")
def test_spreadsheet_to_csvs_xls_failure(mock_open_workbook: MagicMock, converter_service: ConverterService) -> None:
    """Tests that an exception during XLS to CSV conversion is handled."""
    mock_open_workbook.side_effect = Exception("xlrd error")
    with pytest.raises(Exception, match="xlrd error"):
        converter_service.spreadsheet_to_csvs(b"bad xls content", ".xls")


@patch("openpyxl.load_workbook")
def test_spreadsheet_to_csvs_xlsx_failure(mock_load_workbook: MagicMock, converter_service: ConverterService) -> None:
    """Tests that an exception during XLSX to CSV conversion is handled."""
    mock_load_workbook.side_effect = Exception("openpyxl error")
    with pytest.raises(Exception, match="openpyxl error"):
        converter_service.spreadsheet_to_csvs(b"bad xlsx content", ".xlsx")


@patch("public_detective.services.converter.open_xlsb")
def test_spreadsheet_to_csvs_xlsb_failure(mock_open_workbook: MagicMock, converter_service: ConverterService) -> None:
    """Tests that an exception during XLSB to CSV conversion is handled."""
    mock_open_workbook.side_effect = Exception("pyxlsb error")
    with pytest.raises(Exception, match="pyxlsb error"):
        converter_service.spreadsheet_to_csvs(b"bad xlsb content", ".xlsb")
